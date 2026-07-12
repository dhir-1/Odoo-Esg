import io
import csv
import logging
from typing import Dict, Any, List
from datetime import datetime

from fastapi.responses import StreamingResponse, Response

# openpyxl for Excel sheets
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 1. CSV EXPORT
# ---------------------------------------------------------------------------
def export_to_csv(data: Dict[str, Any], filename: str) -> StreamingResponse:
    """
    Exports report structured data to tabular CSV format.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # 1. Report Header
    writer.writerow(["ECOSPHERE REPORT EXPORT"])
    writer.writerow(["Generated At", datetime.now().isoformat()])
    if "filter_criteria" in data:
        criteria = data["filter_criteria"]
        writer.writerow(["Date From", criteria.get("date_from") or "N/A"])
        writer.writerow(["Date To", criteria.get("date_to") or "N/A"])
        writer.writerow(["Modules", criteria.get("module_requested") or "All"])
    writer.writerow([])  # spacer row

    # 2. Environmental Section
    env = data.get("environmental", {})
    if env:
        writer.writerow(["### ENVIRONMENTAL ###"])
        txs = env.get("carbon_transactions", [])
        if txs:
            writer.writerow(["Carbon Transactions"])
            writer.writerow(["ID", "Source Module", "Quantity", "Calculated CO2e", "Date"])
            for tx in txs:
                writer.writerow([tx["id"], tx["source_module"], tx["quantity"], tx["calculated_co2e"], tx["transaction_date"]])
        goals = env.get("goals", [])
        if goals:
            writer.writerow([])
            writer.writerow(["Environmental Goals"])
            writer.writerow(["ID", "Title", "Target Value", "Current Value", "Progress Status"])
            for g in goals:
                writer.writerow([g["id"], g["title"], g["target_value"], g["current_value"], g["progress_status"]])
        writer.writerow([])

    # 3. Social Section
    soc = data.get("social", {})
    if soc:
        writer.writerow(["### SOCIAL ###"])
        parts = soc.get("csr_participations", [])
        if parts:
            writer.writerow(["CSR Participations"])
            writer.writerow(["ID", "Activity Title", "Points Earned", "Approval Status", "Completion Date"])
            for p in parts:
                writer.writerow([p["id"], p["activity_title"], p["points_earned"], p["approval_status"], p["completion_date"]])
        divs = soc.get("diversity_metrics", [])
        if divs:
            writer.writerow([])
            writer.writerow(["Diversity Metrics"])
            writer.writerow(["Category", "Label", "Count", "Period"])
            for d in divs:
                writer.writerow([d["category"], d["label"], d["count"], d["period"]])
        writer.writerow([])

    # 4. Governance Section
    gov = data.get("governance", {})
    if gov:
        writer.writerow(["### GOVERNANCE ###"])
        audits = gov.get("audits", [])
        if audits:
            writer.writerow(["Audits"])
            writer.writerow(["ID", "Title", "Audit Date", "Status", "Overall Rating"])
            for a in audits:
                writer.writerow([a["id"], a["title"], a["audit_date"], a["status"], a["overall_rating"] or "N/A"])
        issues = gov.get("compliance_issues", [])
        if issues:
            writer.writerow([])
            writer.writerow(["Compliance Issues"])
            writer.writerow(["ID", "Severity", "Description", "Status", "Due Date"])
            for i in issues:
                writer.writerow([i["id"], i["severity"], i["description"], i["status"], i["due_date"]])
        writer.writerow([])

    # 5. Gamification Section
    gam = data.get("gamification", {})
    if gam:
        writer.writerow(["### GAMIFICATION ###"])
        challs = gam.get("challenge_participations", [])
        if challs:
            writer.writerow(["Challenge Participations"])
            writer.writerow(["ID", "Challenge Title", "XP Awarded", "Status"])
            for cp in challs:
                writer.writerow([cp["id"], cp["challenge_title"], cp["xp_awarded"], cp["approval_status"]])
        reds = gam.get("reward_redemptions", [])
        if reds:
            writer.writerow([])
            writer.writerow(["Reward Redemptions"])
            writer.writerow(["ID", "Reward Title", "Points Spent", "Redeemed At"])
            for rr in reds:
                writer.writerow([rr["id"], rr["reward_title"], rr["points_spent"], rr["redeemed_at"]])

    output.seek(0)
    return StreamingResponse(
        io.StringIO(output.getvalue()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# 2. EXCEL (XLSX) EXPORT
# ---------------------------------------------------------------------------
def export_to_xlsx(data: Dict[str, Any], filename: str) -> Response:
    """
    Exports report structured data to designed Excel file with openpyxl.
    Uses multi-sheet tabs for each module and styles them nicely.
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    def style_header_row(ws, max_col):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_fit_columns(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 1. Environmental Sheet
    env = data.get("environmental", {})
    if env:
        ws = wb.create_sheet(title="Environmental")
        ws.append(["ID", "Type", "Source Module", "Quantity", "Calculated CO2e", "Date / Target Status"])
        style_header_row(ws, 6)

        row_num = 2
        txs = env.get("carbon_transactions", [])
        for tx in txs:
            ws.append([tx["id"], "Emission Transaction", tx["source_module"], tx["quantity"], tx["calculated_co2e"], tx["transaction_date"]])
            row_num += 1
        
        goals = env.get("goals", [])
        for g in goals:
            ws.append([g["id"], "Environmental Goal", g["title"], g["target_value"], g["current_value"], g["progress_status"]])
            row_num += 1

        for r in range(2, row_num):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
        auto_fit_columns(ws)

    # 2. Social Sheet
    soc = data.get("social", {})
    if soc:
        ws = wb.create_sheet(title="Social")
        ws.append(["ID/Ref", "Type", "Title/Category", "Metric 1", "Metric 2", "Status/Period"])
        style_header_row(ws, 6)

        row_num = 2
        parts = soc.get("csr_participations", [])
        for p in parts:
            ws.append([p["id"], "CSR Participation", p["activity_title"], p["points_earned"], "", p["approval_status"]])
            row_num += 1
        
        divs = soc.get("diversity_metrics", [])
        for d in divs:
            ws.append(["", "Diversity Metric", d["category"], d["label"], d["count"], d["period"]])
            row_num += 1

        for r in range(2, row_num):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
        auto_fit_columns(ws)

    # 3. Governance Sheet
    gov = data.get("governance", {})
    if gov:
        ws = wb.create_sheet(title="Governance")
        ws.append(["ID", "Type", "Description/Title", "Audit Date/Category", "Rating/Severity", "Status"])
        style_header_row(ws, 6)

        row_num = 2
        audits = gov.get("audits", [])
        for a in audits:
            ws.append([a["id"], "Audit", a["title"], a["audit_date"], a["overall_rating"] or "", a["status"]])
            row_num += 1
        
        issues = gov.get("compliance_issues", [])
        for i in issues:
            ws.append([i["id"], "Compliance Issue", i["description"], "", i["severity"], i["status"]])
            row_num += 1

        for r in range(2, row_num):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
        auto_fit_columns(ws)

    # 4. Gamification Sheet
    gam = data.get("gamification", {})
    if gam:
        ws = wb.create_sheet(title="Gamification")
        ws.append(["ID", "Type", "Title", "Points/XP", "Redemption Date", "Status"])
        style_header_row(ws, 6)

        row_num = 2
        challs = gam.get("challenge_participations", [])
        for cp in challs:
            ws.append([cp["id"], "Challenge Completion", cp["challenge_title"], cp["xp_awarded"], "", cp["approval_status"]])
            row_num += 1
        
        reds = gam.get("reward_redemptions", [])
        for rr in reds:
            ws.append([rr["id"], "Reward Redemption", rr["reward_title"], rr["points_spent"], rr["redeemed_at"], ""])
            row_num += 1

        for r in range(2, row_num):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.font = regular_font
                cell.border = thin_border
        auto_fit_columns(ws)

    # If no sheets were created, create a default overview
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Report Overview")
        ws.append(["Report contains no data for the requested filters."])
        ws.cell(row=1, column=1).font = bold_font

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# 3. PDF EXPORT
# ---------------------------------------------------------------------------
def export_to_pdf(data: Dict[str, Any], filename: str) -> Response:
    """
    Exports report structured data to a styled PDF document using ReportLab.
    ReportLab is chosen because it is pure python and does not require complex
    external native command-line dependencies (unlike weasyprint, which requires
    installing Pango/Cairo binaries that frequently fail or mismatch on Windows setups).
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=12
    )

    section_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1F4E79'),
        spaceBefore=14,
        spaceAfter=6
    )

    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#595959')
    )

    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )

    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    story = []

    # Title
    story.append(Paragraph("EcoSphere ESG Performance Report", title_style))
    
    # Metadata / Filter Details
    meta_lines = [
        f"<b>Generated At:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ]
    if "filter_criteria" in data:
        crit = data["filter_criteria"]
        meta_lines.append(f"<b>Date Period:</b> {crit.get('date_from') or 'Start'} to {crit.get('date_to') or 'End'}")
        meta_lines.append(f"<b>Module Filtered:</b> {crit.get('module_requested') or 'All Modules'}")
    
    for line in meta_lines:
        story.append(Paragraph(line, meta_style))
    story.append(Spacer(1, 15))

    # Helper table styler
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F2F2F2'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # 1. Environmental Section
    env = data.get("environmental", {})
    if env:
        story.append(Paragraph("Environmental Metrics", section_style))
        txs = env.get("carbon_transactions", [])
        if txs:
            table_data = [[
                Paragraph("Source Module", header_cell_style),
                Paragraph("Quantity", header_cell_style),
                Paragraph("CO2e Emitted", header_cell_style),
                Paragraph("Transaction Date", header_cell_style)
            ]]
            for tx in txs[:10]:  # Limit rows in PDF to keep it tidy
                table_data.append([
                    Paragraph(tx["source_module"], body_style),
                    Paragraph(str(tx["quantity"]), body_style),
                    Paragraph(str(tx["calculated_co2e"]), body_style),
                    Paragraph(tx["transaction_date"], body_style)
                ])
            t = Table(table_data, colWidths=[130, 130, 130, 130])
            t.setStyle(table_style)
            story.append(t)
        else:
            story.append(Paragraph("No emissions transactions found for this scope.", body_style))
        story.append(Spacer(1, 10))

    # 2. Social Section
    soc = data.get("social", {})
    if soc:
        story.append(Paragraph("Social Metrics", section_style))
        divs = soc.get("diversity_metrics", [])
        if divs:
            table_data = [[
                Paragraph("Diversity Category", header_cell_style),
                Paragraph("Label / Segment", header_cell_style),
                Paragraph("Headcount", header_cell_style),
                Paragraph("Reporting Period", header_cell_style)
            ]]
            for d in divs[:10]:
                table_data.append([
                    Paragraph(d["category"], body_style),
                    Paragraph(d["label"], body_style),
                    Paragraph(str(d["count"]), body_style),
                    Paragraph(d["period"], body_style)
                ])
            t = Table(table_data, colWidths=[130, 130, 130, 130])
            t.setStyle(table_style)
            story.append(t)
        else:
            story.append(Paragraph("No diversity or CSR participation records found for this scope.", body_style))
        story.append(Spacer(1, 10))

    # 3. Governance Section
    gov = data.get("governance", {})
    if gov:
        story.append(Paragraph("Governance & Risk Metrics", section_style))
        audits = gov.get("audits", [])
        if audits:
            table_data = [[
                Paragraph("Audit Title", header_cell_style),
                Paragraph("Date Conducted", header_cell_style),
                Paragraph("Overall Rating", header_cell_style),
                Paragraph("Current Status", header_cell_style)
            ]]
            for a in audits[:10]:
                table_data.append([
                    Paragraph(a["title"], body_style),
                    Paragraph(a["audit_date"], body_style),
                    Paragraph(str(a["overall_rating"]) if a["overall_rating"] else "N/A", body_style),
                    Paragraph(a["status"], body_style)
                ])
            t = Table(table_data, colWidths=[180, 110, 110, 120])
            t.setStyle(table_style)
            story.append(t)
        else:
            story.append(Paragraph("No internal audit results found for this scope.", body_style))
        story.append(Spacer(1, 10))

    # Build PDF
    doc.build(story)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
